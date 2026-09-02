import os
from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, Request, Form, HTTPException, Query
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

from database import init_db, get_db
from whatsapp import (send_whatsapp_message, send_quantity_list, send_area_list,
                      send_time_list, send_date_list, send_site_list,
                      download_whatsapp_media, send_order_confirmation_card,
                      send_contact_buttons, time_greeting)
from scheduler import start_scheduler, reschedule, reschedule_admin

AREA_DISPLAY = {
    "area_jerusalem":    "ירושלים והסביבה",
    "area_modiin":       "מודיעין",
    "area_maale_adumim": "מעלה אדומים",
    "area_beit_shemesh": "בית שמש",
    "area_other":        "אחר",
}
AREA_DRIVER_KEY = {
    "area_jerusalem":    "ירושלים",
    "area_modiin":       "מודיעין",
    "area_maale_adumim": "מעלה אדומים",
    "area_beit_shemesh": "בית שמש",
    "area_other":        "אחר",
}
QUANTITY_MAP = {
    "qty_0_100":     "0–100",
    "qty_100_400":   "100–400",
    "qty_400_1000":  "400–1,000",
    "qty_1000_2000": "1,000–2,000",
}

# מניעת עיבוד כפול של אותו webhook
_processed_msg_ids: set = set()
TIME_MAP = {
    "time_07_09": "07:00–09:00",
    "time_09_11": "09:00–11:00",
    "time_11_13": "11:00–13:00",
    "time_13_15": "13:00–15:00",
}

def _parse_quantity_lower(quantity_str: str) -> int:
    """מחזיר את הגבול התחתון של הכמות כמספר שלם."""
    try:
        return int(quantity_str.split("–")[0].replace(",", "").strip())
    except Exception:
        return 0


def _tanker_vol(driver) -> int:
    try:
        return int(str(driver["tanker_volume"]).replace(",", ""))
    except Exception:
        return 0


def assign_driver(conn, area: str, quantity_str: str):
    """
    מחזיר נהג מתאים לפי אזור וכמות.
    כלל מיוחד: ירושלים + מעל 400 ליטר → מעדיף את 'ישראל מזרחי'.
    """
    qty_lower = _parse_quantity_lower(quantity_str)
    drivers   = conn.execute("SELECT * FROM drivers").fetchall()

    # סנן נהגים שפועלים באזור הנדרש
    def serves_area(d):
        return any(area.strip() in a.strip() for a in d["area"].split(","))

    candidates = [d for d in drivers if serves_area(d)]
    if not candidates:
        candidates = drivers  # fallback: כל הנהגים

    # כלל מיוחד: ירושלים + מעל 400 ליטר → ישראל מזרחי
    if "ירושלים" in area and qty_lower >= 400:
        israel = next((d for d in candidates if "ישראל מזרחי" in d["name"]), None)
        if israel:
            return israel

    # העדף נהג שמיכליתו מספיקה לכמות — הקטנה שמתאימה (יעילות)
    fitting = [d for d in candidates if _tanker_vol(d) >= qty_lower]
    if fitting:
        return min(fitting, key=_tanker_vol)

    # אם אין — קח בעל המיכלית הגדולה ביותר
    return max(candidates, key=_tanker_vol) if candidates else None


from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="מערכת הזמנות סולר")
_ORIGINS = [o.strip() for o in os.getenv("CORS_ORIGINS", "*").split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ORIGINS,
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static", StaticFiles(directory="static"), name="static")

BASE_URL = os.getenv("BASE_URL", "http://localhost:8000")
VERIFY_TOKEN = os.getenv("WHATSAPP_VERIFY_TOKEN", "secret")


# ── Startup ──────────────────────────────────────────────────────────────────

@app.on_event("startup")
def startup():
    init_db()
    with get_db() as conn:
        settings = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
    hour   = int(settings.get("daily_hour", 14))
    minute = int(settings.get("daily_minute", 0))
    admin_hour   = int(settings["admin_schedule_hour"])   if settings.get("admin_schedule_hour")   else None
    admin_minute = int(settings.get("admin_schedule_minute", 0))
    start_scheduler(hour, minute, admin_hour, admin_minute)
    from recurring import materialize_recurring_orders
    from datetime import date as _d, timedelta as _td
    _tomorrow = (_d.today() + _td(days=1)).isoformat()
    _n = materialize_recurring_orders(_tomorrow)
    if _n:
        print(f"[Startup] נוצרו {_n} הזמנות קבועות ל-{_tomorrow}")


# ── Order document parsing ───────────────────────────────────────────────────

def _extract_pdf_text(file_bytes: bytes) -> str:
    try:
        import pdfplumber, io
        parts = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    parts.append(t)
        return "\n".join(parts)
    except Exception as e:
        print(f"[PDF] {e}")
        return ""


def _regex_parse(text: str) -> dict:
    import re
    from datetime import datetime
    result = {"quantity": None, "delivery_date": None, "address": None,
              "contact_name": None, "contact_phone": None}
    qty = re.search(r'(\d[\d,]*)\s*(?:ל[יי][טת]ר|ל[\'"]|lit)', text, re.IGNORECASE)
    if qty:
        result["quantity"] = int(qty.group(1).replace(",", ""))
    dt = re.search(r'(\d{1,2})[/\.](\d{1,2})[/\.](\d{2,4})', text)
    if dt:
        d, m, y = dt.groups()
        if len(y) == 2: y = "20" + y
        try:
            result["delivery_date"] = datetime(int(y), int(m), int(d)).strftime("%Y-%m-%d")
        except Exception:
            pass
    phone = re.search(r'(0\d[-\s]?\d{3}[-\s]?\d{4})', text)
    if phone:
        result["contact_phone"] = phone.group(1)
    addr = re.search(r'(?:כתובת|משלוח ל|אתר)[:\s]+([^\n]{5,60})', text, re.IGNORECASE)
    if addr:
        result["address"] = addr.group(1).strip()
    return result


def parse_order_document(file_bytes: bytes, mime_type: str, customer_name: str) -> dict | None:
    import json, base64, re as _re
    text_content = ""
    if "pdf" in mime_type.lower():
        text_content = _extract_pdf_text(file_bytes)

    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if api_key:
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)
            system = (
                'חלץ פרטי הזמנת דלק סולר. החזר JSON בלבד:\n'
                '{"quantity":<ליטרים כמספר שלם>,"delivery_date":"<YYYY-MM-DD או null>",'
                '"address":"<כתובת מלאה או null>","contact_name":"<שם או null>","contact_phone":"<טלפון או null>"}'
            )
            if text_content:
                msgs = [{"role": "user", "content": f"לקוח: {customer_name}\n\n{text_content[:3000]}"}]
            else:
                for mt in ["image/jpeg", "image/png", "image/webp"]:
                    if mt in mime_type:
                        real_mt = mt; break
                else:
                    real_mt = "image/jpeg"
                img_b64 = base64.standard_b64encode(file_bytes).decode()
                msgs = [{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": real_mt, "data": img_b64}},
                    {"type": "text", "text": f"לקוח: {customer_name}\nחלץ פרטי הזמנה"},
                ]}]
            resp = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=400,
                                          system=system, messages=msgs)
            raw = resp.content[0].text
            m = _re.search(r'\{.*?\}', raw, _re.DOTALL)
            if m:
                return json.loads(m.group())
        except Exception as e:
            print(f"[Claude parse] {e}")

    if text_content:
        return _regex_parse(text_content)
    return None


# ── WhatsApp Webhook ──────────────────────────────────────────────────────────

@app.get("/webhook")
def verify_webhook(request: Request):
    params = dict(request.query_params)
    if params.get("hub.verify_token") == VERIFY_TOKEN:
        return int(params.get("hub.challenge", 0))
    raise HTTPException(status_code=403, detail="Forbidden")


def _fmt_phone(p: str) -> str:
    p = (p or "").strip().replace(" ", "").replace("-", "").lstrip("+")
    if p.startswith("0"):
        p = "972" + p[1:]
    return p


def _notify_admin_new_order(order_info: dict):
    """שולח התרעה למנהל בוואטסאפ על כל הזמנה חדשה."""
    try:
        with get_db() as conn:
            settings = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
        admin_phone = settings.get("admin_phone", "").strip().replace(" ", "").replace("-", "").lstrip("+")
        if not admin_phone:
            return
        if admin_phone.startswith("0"):
            admin_phone = "972" + admin_phone[1:]
        customer = order_info.get("customer_name", "לא ידוע")
        address  = order_info.get("site_address", "")
        qty      = order_info.get("quantity", "")
        date     = order_info.get("order_date", "")
        driver   = order_info.get("driver_name", "")
        msg = f"🔔 *הזמנה חדשה נכנסה!*\n\n👥 {customer}\n📍 {address}\n⛽ {qty} ליטר"
        if date:
            try:
                from datetime import datetime as _dt2
                d = _dt2.strptime(date, "%Y-%m-%d")
                msg += f"\n📅 {d.day}/{d.month}/{d.year}"
            except Exception:
                msg += f"\n📅 {date}"
        if driver:
            msg += f"\n🚛 נהג: {driver}"
        send_whatsapp_message(admin_phone, msg)
    except Exception as e:
        print(f"[Notify] שגיאה בשליחת התרעה למנהל: {e}")


def _send_order_immediate(driver: dict, order_info: dict) -> dict:
    """שולח הזמנה מיידית לנהג ולמנהל (להעברה לקבוצה). מחזיר סטטוס שליחה."""
    from datetime import datetime as _dt3
    date_str = order_info.get("order_date", "")
    try:
        d = _dt3.strptime(date_str, "%Y-%m-%d")
        date_label = f"{d.day}/{d.month}/{d.year}"
    except Exception:
        date_label = date_str or "היום"

    cname  = order_info.get("contact_name", "")
    cphone = order_info.get("contact_phone", "")
    contact_line = f"\n📞 {cname}" + (f" — {cphone}" if cphone else "") if (cname or cphone) else ""

    driver_msg = (
        f"📋 *הזמנה מיידית — {date_label}*\n\n"
        f"👥 {order_info['customer_name']}\n"
        f"📍 {order_info['site_address']}\n"
        f"⛽ {order_info['quantity']} ליטר"
        + contact_line
    )

    target = _fmt_phone(driver.get("personal_phone") or "") or _fmt_phone(driver.get("phone") or "")
    driver_ok = bool(target) and send_whatsapp_message(target, driver_msg)

    admin_ok = False
    try:
        with get_db() as conn:
            settings = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
        admin_phone = _fmt_phone(settings.get("admin_phone", ""))
        if admin_phone:
            admin_msg = (
                f"📦 *הזמנה מיידית — {date_label}*\n\n"
                f"👥 *{order_info['customer_name']}*\n"
                f"📍 {order_info['site_address']}\n"
                f"⛽ {order_info['quantity']} ליטר"
                + contact_line +
                f"\n🚛 נהג: {driver['name']}\n\n"
                f"✅ נשלחה לנהג"
            )
            admin_ok = send_whatsapp_message(admin_phone, admin_msg)
    except Exception as e:
        print(f"[ImmediateSend] שגיאה בשליחה למנהל: {e}")

    return {
        "driver_ok": bool(driver_ok),
        "admin_ok": bool(admin_ok),
        "driver_name": driver["name"],
        "driver_phone": target,
    }


@app.post("/webhook")
async def receive_message(request: Request):
    data = await request.json()

    # סטטוס קריאה
    try:
        for s in data["entry"][0]["changes"][0]["value"].get("statuses", []):
            if s.get("status") == "read":
                with get_db() as conn:
                    conn.execute("UPDATE customers SET msg_read=1 WHERE last_msg_id=?", (s["id"],))
    except Exception:
        pass

    try:
        entry   = data["entry"][0]["changes"][0]["value"]
        message = entry["messages"][0]
        msg_id  = message.get("id", "")
        phone   = message["from"]
        msg_type = message.get("type", "text")

        # מניעת עיבוד כפול
        if msg_id and msg_id in _processed_msg_ids:
            return JSONResponse({"status": "ok"})
        if msg_id:
            _processed_msg_ids.add(msg_id)
            if len(_processed_msg_ids) > 500:
                _processed_msg_ids.clear()

        # ── פקודות מנהל ──────────────────────────────────────────────────────
        if msg_type == "text":
            _raw_text = message["text"]["body"].strip()
            if _raw_text in ("סידור", "שלח סידור", "schedule"):
                with get_db() as conn:
                    _s = {r["key"]: r["value"] for r in conn.execute("SELECT key,value FROM settings").fetchall()}
                _admin_p = _s.get("admin_phone", "").strip().replace(" ", "").replace("-", "")
                if _admin_p.startswith("0"):
                    _admin_p = "972" + _admin_p[1:]
                _phone_alt = ("0" + phone[3:]) if phone.startswith("972") else ("972" + phone[1:])
                if phone == _admin_p or _phone_alt == _admin_p:
                    from datetime import date as _d, timedelta as _td
                    from whatsapp import send_order_card as _soc
                    _target = (_d.today() + _td(days=1)).isoformat()
                    with get_db() as conn:
                        _orders = conn.execute("""
                            SELECT o.*, d.name as driver_name, d.phone as driver_phone,
                                   d.personal_phone as driver_personal_phone
                            FROM orders o LEFT JOIN drivers d ON o.driver_id=d.id
                            WHERE o.order_date=?
                            ORDER BY d.name, o.sort_order, o.delivery_time, o.created_at
                        """, (_target,)).fetchall()
                    _orders = [dict(o) for o in _orders]
                    if not _orders:
                        send_whatsapp_message(phone, f"📋 סידור יומי — {_target}\n\nאין הזמנות למחר.")
                    else:
                        _by_drv = {}
                        for _o in _orders:
                            _by_drv.setdefault(_o["driver_name"] or "ללא נהג", []).append(_o)
                        _lines = [f"📋 *{_target}* — {len(_orders)} הזמנות\n"]
                        for _dn, _dos in _by_drv.items():
                            _lines.append(f"*{_dn}:*")
                            for _i, _o in enumerate(_dos, 1):
                                _lines.append(f"{_i}. {_o['customer_name']} — {_o['site_address']}")
                            _lines.append("")
                        send_whatsapp_message(phone, "\n".join(_lines))
                    return JSONResponse({"status": "ok"})

        # כפתור ביצוע נהג — לפני כל בדיקה אחרת
        if msg_type == "interactive":
            inter = message["interactive"]
            if inter.get("type") == "button_reply":
                btn_id = inter["button_reply"]["id"]
                if btn_id.startswith("done_"):
                    import json as _json_done
                    order_id = int(btn_id.split("_")[1])
                    _pa = ("0" + phone[3:]) if phone.startswith("972") else ("972" + phone[1:])
                    with get_db() as conn:
                        _is_drv = conn.execute(
                            "SELECT id FROM drivers WHERE phone=? OR phone=? OR personal_phone=? OR personal_phone=?",
                            (phone, _pa, phone, _pa)
                        ).fetchone()
                        _ord = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
                        if _is_drv:
                            # שמור מצב "ממתין לכמות בפועל"
                            conn.execute(
                                "INSERT OR REPLACE INTO conversation_state "
                                "(phone, step, pending_order_json, updated_at) "
                                "VALUES (?, 'driver_awaiting_qty', ?, datetime('now','localtime'))",
                                (phone, _json_done.dumps({"order_id": order_id}))
                            )
                        else:
                            conn.execute("UPDATE orders SET status='הושלם' WHERE id=?", (order_id,))
                    if _is_drv and _ord:
                        send_whatsapp_message(phone,
                            f"📦 הזמנה #{order_id} — *{_ord['customer_name']}*\n"
                            f"כמות מוזמנת: *{_ord['quantity']} ליטר*\n\n"
                            f"כמה ליטרים סיפקת בפועל? (הזן מספר)"
                        )
                    else:
                        send_whatsapp_message(phone, f"✅ הזמנה #{order_id} סומנה כהושלם!")
                    return JSONResponse({"status": "ok"})

        if msg_type == "text":
            text     = message["text"]["body"].strip()
            reply_id = None
        elif msg_type == "interactive":
            inter = message["interactive"]
            if inter["type"] == "list_reply":
                reply_id = inter["list_reply"]["id"]
                text     = inter["list_reply"]["title"]
            elif inter["type"] == "button_reply":
                reply_id = inter["button_reply"]["id"]
                text     = inter["button_reply"]["title"]
            else:
                return JSONResponse({"status": "ok"})
        elif msg_type in ("document", "image"):
            text = ""; reply_id = None
        else:
            return JSONResponse({"status": "ok"})
    except (KeyError, IndexError):
        return JSONResponse({"status": "ok"})

    # ── זהה נהג ──────────────────────────────────────────────────────────────
    phone_alt = ("0" + phone[3:]) if phone.startswith("972") else ("972" + phone[1:])
    with get_db() as conn:
        driver_row = conn.execute(
            "SELECT * FROM drivers WHERE phone=? OR phone=? OR personal_phone=? OR personal_phone=?",
            (phone, phone_alt, phone, phone_alt)
        ).fetchone()

    if driver_row:
        import json as _json_drv
        drv = dict(driver_row)

        # בדוק אם הנהג ממתין לאישור כמות
        with get_db() as conn:
            _drv_state = conn.execute(
                "SELECT * FROM conversation_state WHERE phone=?", (phone,)
            ).fetchone()

        if _drv_state and _drv_state["step"] == "driver_awaiting_qty" and msg_type == "text":
            _info = _json_drv.loads(_drv_state["pending_order_json"] or "{}")
            _oid  = _info.get("order_id")
            _qty_text = text.strip()
            if _oid and _qty_text:
                from datetime import datetime as _dt
                import pytz as _pytz
                _now_il = _dt.now(_pytz.timezone("Asia/Jerusalem"))
                _delivery_time = _now_il.strftime("%H:%M")
                with get_db() as conn:
                    conn.execute(
                        "UPDATE orders SET status='הושלם', actual_quantity=?, delivery_time=? WHERE id=?",
                        (_qty_text, _delivery_time, _oid)
                    )
                    conn.execute("DELETE FROM conversation_state WHERE phone=?", (phone,))
                send_whatsapp_message(phone, f"✅ הזמנה #{_oid} הושלמה — *{_qty_text} ליטר* נקלטו בשעה {_delivery_time}. תודה!")
            return JSONResponse({"status": "ok"})

        # נהג ביקש את הסידור שלו
        if msg_type == "text" and text.strip() in ("סידור", "שלח סידור", "schedule"):
            from datetime import date as _d2, timedelta as _td2
            from whatsapp import send_order_card as _soc2
            _target2 = (_d2.today() + _td2(days=1)).isoformat()
            with get_db() as conn:
                _drv_orders = conn.execute("""
                    SELECT o.* FROM orders o
                    WHERE o.driver_id=? AND o.order_date=?
                    ORDER BY o.sort_order, o.delivery_time, o.created_at
                """, (drv["id"], _target2)).fetchall()
            _drv_orders = [dict(o) for o in _drv_orders]
            if not _drv_orders:
                send_whatsapp_message(phone, f"📋 {_target2}\nאין לך הזמנות למחר.")
            else:
                send_whatsapp_message(phone, f"📋 *סידור יומי — {_target2}*\nיש לך {len(_drv_orders)} הזמנות:")
                for _i2, _o2 in enumerate(_drv_orders, 1):
                    _soc2(phone, _o2, _i2, len(_drv_orders))
        # כל שאר ההודעות מנהג — מתעלמים (לא מוסיפים לממתינים)
        return JSONResponse({"status": "ok"})

    # מצא לקוח (תומך 0XXXXXXXX ו-972XXXXXXXX, phone/phone2/contact_phone/order_contact_phone/customer_contacts)
    with get_db() as conn:
        customer = conn.execute("""
            SELECT * FROM customers WHERE active=1 AND (
                phone=? OR phone=? OR phone2=? OR phone2=?
                OR contact_phone=? OR contact_phone=?
                OR order_contact_phone=? OR order_contact_phone=?
            )
            UNION
            SELECT c.* FROM customers c
            JOIN customer_contacts cc ON cc.customer_id=c.id
            WHERE cc.active=1 AND c.active=1
              AND (cc.phone=? OR cc.phone=?)
            LIMIT 1
        """, (phone, phone_alt, phone, phone_alt,
              phone, phone_alt, phone, phone_alt,
              phone, phone_alt)
        ).fetchone()
        state = conn.execute(
            "SELECT * FROM conversation_state WHERE phone=?", (phone,)
        ).fetchone()

    if not customer:
        with get_db() as conn:
            pend = conn.execute("SELECT * FROM pending_registrations WHERE phone=?", (phone,)).fetchone()
            pend_state = conn.execute("SELECT * FROM conversation_state WHERE phone=? AND step='pending_awaiting_company'", (phone,)).fetchone()
        if pend_state:
            # לקוח לא מוכר השיב עם שם חברה
            company_name = text.strip()
            with get_db() as conn:
                conn.execute(
                    "UPDATE pending_registrations SET company_name=?, status='pending' WHERE phone=?",
                    (company_name, phone)
                )
                conn.execute("DELETE FROM conversation_state WHERE phone=?", (phone,))
            send_whatsapp_message(phone, "תודה! נציג מאיתנו יצור איתך קשר בקרוב 😊")
        elif not pend:
            # מספר חדש לחלוטין — שמור ושאל שם חברה
            with get_db() as conn:
                conn.execute(
                    "INSERT OR IGNORE INTO pending_registrations (phone) VALUES (?)", (phone,)
                )
                conn.execute(
                    "INSERT OR REPLACE INTO conversation_state (phone, step) VALUES (?, 'pending_awaiting_company')",
                    (phone,)
                )
            send_whatsapp_message(
                phone,
                "שלום! אנחנו *וזאת הברכה דלקים* 🛢️\n"
                "מספרך אינו רשום במערכת.\n\n"
                "מה שם החברה / העסק שלך?"
            )
        else:
            send_whatsapp_message(phone, "תודה, נציג יצור איתך קשר בקרוב 😊")
        return JSONResponse({"status": "ok"})

    step = state["step"] if state else None

    # טיפול בקובץ PDF / תמונה — הזמנת רכש
    if msg_type in ("document", "image"):
        import json as _json
        media_obj = message.get(msg_type, {})
        media_id  = media_obj.get("id")
        if not media_id:
            return JSONResponse({"status": "ok"})
        send_whatsapp_message(phone, "⏳ מעבד את ההזמנה, רגע...")
        extracted_text = ""
        try:
            file_bytes, mime_type = download_whatsapp_media(media_id)
            print(f"[Doc] mime={mime_type} size={len(file_bytes)}")
            if "pdf" in mime_type.lower():
                extracted_text = _extract_pdf_text(file_bytes)
            print(f"[Doc] text={repr(extracted_text[:300])}")
            info = parse_order_document(file_bytes, mime_type, customer["name"])
            print(f"[Doc] parsed={info}")
        except Exception as e:
            print(f"[Doc handler] {e}")
            info = None
        if not info or not info.get("quantity"):
            # שלח טקסט שחולץ כדי לדעת מה יש בקובץ
            debug_msg = f"לא הצלחתי לחלץ כמות.\nטקסט שנקרא:\n{extracted_text[:400] if extracted_text else '(ריק — PDF סרוק?)'}"
            send_whatsapp_message(phone, debug_msg)
            return JSONResponse({"status": "ok"})
        # fill missing contact from customer
        if not info.get("contact_name"):  info["contact_name"]  = customer["contact_name"]  or ""
        if not info.get("contact_phone"): info["contact_phone"] = customer["contact_phone"] or ""
        if not info.get("address"):       info["address"]       = customer["site_address"]  or ""
        with get_db() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO conversation_state (phone, step, customer_id, pending_order_json) VALUES (?,?,?,?)",
                (phone, "awaiting_order_confirmation", customer["id"], _json.dumps(info, ensure_ascii=False))
            )
        send_order_confirmation_card(phone, info)
        return JSONResponse({"status": "ok"})

    # אישור / ביטול הזמנה מקובץ
    if step == "awaiting_order_confirmation":
        import json as _json
        if reply_id == "confirm_order":
            raw = state["pending_order_json"] if state else None
            if raw:
                info = _json.loads(raw)
                from datetime import date as _date, timedelta
                order_date = info.get("delivery_date") or (_date.today() + timedelta(days=1)).isoformat()
                qty_str    = str(info.get("quantity", "0–100"))
                address    = info.get("address", "")
                cname      = info.get("contact_name")  or customer["contact_name"]  or ""
                cphone     = info.get("contact_phone") or customer["contact_phone"] or ""
                city       = customer["area"] or "ירושלים"
                with get_db() as conn:
                    driver = assign_driver(conn, city, qty_str)
                    driver_id = driver["id"] if driver else None
                    conn.execute(
                        """INSERT INTO orders
                           (customer_id,customer_name,site_address,contact_name,contact_phone,
                            quantity,driver_id,order_date,delivery_time)
                           VALUES (?,?,?,?,?,?,?,?,?)""",
                        (customer["id"], customer["name"], address, cname, cphone,
                         qty_str, driver_id, order_date, "")
                    )
                    conn.execute("DELETE FROM conversation_state WHERE phone=?", (phone,))
                _notify_admin_new_order({
                    "customer_name": customer["name"], "site_address": address,
                    "quantity": qty_str, "order_date": order_date,
                    "driver_name": driver["name"] if driver else "",
                })
                from datetime import datetime as _dt
                try:
                    d = _dt.strptime(order_date, "%Y-%m-%d")
                    date_label = f"{d.day}/{d.month}/{d.year}"
                except Exception:
                    date_label = order_date
                send_whatsapp_message(phone,
                    f"ההזמנה אושרה! ✅\n\nכמות: {qty_str} ליטר\nתאריך: {date_label}\nכתובת: {address}\nנדאג לאספקה. תודה! 🙏")
            return JSONResponse({"status": "ok"})
        if reply_id == "cancel_order":
            with get_db() as conn:
                conn.execute("DELETE FROM conversation_state WHERE phone=?", (phone,))
            send_whatsapp_message(phone, "ההזמנה בוטלה.")
            return JSONResponse({"status": "ok"})
        return JSONResponse({"status": "ok"})

    # זיהוי כוונת הזמנה עצמאית ("אני רוצה הזמנה למחר" וכד')
    ORDER_KEYWORDS = ["הזמנה", "להזמין", "רוצה הזמנה", "בצע הזמנה", "סולר", "דלק"]
    text_lower = text.strip().lower()
    _not_simple = text_lower not in {"כן", "yes", "1", "כן!", "כן.", "כן,", "אישור", "לא", "no"}
    is_order_intent = any(kw in text_lower for kw in ORDER_KEYWORDS) and _not_simple
    if is_order_intent:
        with get_db() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO conversation_state (phone, step, customer_id) VALUES (?,?,?)",
                (phone, "awaiting_date", customer["id"])
            )
        send_date_list(phone)
        return JSONResponse({"status": "ok"})

    # שלב awaiting_date — לקוח בחר תאריך
    if step == "awaiting_date":
        from datetime import date as dt, timedelta
        if reply_id == "date_immediate":
            order_date = dt.today().isoformat()
            date_label = "היום — משלוח מיידי עד 3 שעות ⚡"
        elif reply_id and reply_id.startswith("date_"):
            order_date = reply_id[5:]
            from datetime import datetime
            d = datetime.strptime(order_date, "%Y-%m-%d")
            date_label = f"{d.day}/{d.month}/{d.year}"
        else:
            order_date = (dt.today() + timedelta(days=1)).isoformat()
            from datetime import datetime
            d = datetime.strptime(order_date, "%Y-%m-%d")
            date_label = f"{d.day}/{d.month}/{d.year}"
        with get_db() as conn:
            conn.execute(
                "UPDATE conversation_state SET step=?, order_date=?, updated_at=datetime('now','localtime') WHERE phone=?",
                ("awaiting_area", order_date, phone)
            )
        send_whatsapp_message(phone, f"מצוין! תאריך אספקה: {date_label} 📅")
        send_area_list(phone)
        return JSONResponse({"status": "ok"})

    # שלב 0 — לקוח ענה "לא" להודעה היומית
    negative = {"לא", "no", "0", "לא.", "לא,", "לא!"}
    if step is None and text_lower in {r.lower() for r in negative}:
        greeting = time_greeting()
        send_whatsapp_message(phone, f"תודה רבה והמשך {greeting} 😊")
        return JSONResponse({"status": "ok"})

    # שלב 0 — לקוח ענה "כן" להודעה היומית
    positive = {"כן", "yes", "1", "כן!", "כן.", "כן,", "אישור"}
    if step is None and text_lower in {r.lower() for r in positive}:
        from datetime import date as dt, timedelta
        tomorrow = (dt.today() + timedelta(days=1)).isoformat()
        with get_db() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO conversation_state (phone, step, customer_id, order_date) VALUES (?,?,?,?)",
                (phone, "awaiting_area", customer["id"], tomorrow)
            )
        send_whatsapp_message(phone, f"אוקיי {customer['name']} בוא נתחיל בהזמנה 😊")
        send_area_list(phone)
        return JSONResponse({"status": "ok"})

    # שלב 1 — בחירת אזור
    if step == "awaiting_area":
        if text_lower in {r.lower() for r in positive}:
            # שלח "כן" שוב — מתחילים מחדש
            with get_db() as conn:
                conn.execute("DELETE FROM conversation_state WHERE phone=?", (phone,))
            send_whatsapp_message(phone, f"אוקיי {customer['name']} בוא נתחיל בהזמנה 😊")
            send_area_list(phone)
            return JSONResponse({"status": "ok"})
        if reply_id not in AREA_DISPLAY:
            send_whatsapp_message(phone, "אנא בחר אזור מהרשימה 👇")
            send_area_list(phone)
            return JSONResponse({"status": "ok"})
        city_display = AREA_DISPLAY[reply_id]
        driver_key   = AREA_DRIVER_KEY[reply_id]
        if reply_id == "area_other":
            with get_db() as conn:
                conn.execute(
                    "UPDATE conversation_state SET step=?, city=?, updated_at=datetime('now','localtime') WHERE phone=?",
                    ("awaiting_custom_area", driver_key, phone)
                )
            send_whatsapp_message(phone, "אוקיי, באיזה אזור תרצה את ההזמנה? 📍")
            return JSONResponse({"status": "ok"})
        with get_db() as conn:
            conn.execute(
                "UPDATE conversation_state SET step=?, city=?, updated_at=datetime('now','localtime') WHERE phone=?",
                ("awaiting_quantity", driver_key, phone)
            )
        send_quantity_list(phone)
        return JSONResponse({"status": "ok"})

    # שלב 1.5 — הזנת אזור חופשי (לאחר בחירת "אחר")
    if step == "awaiting_custom_area":
        custom_area = text.strip()
        with get_db() as conn:
            conn.execute(
                "UPDATE conversation_state SET step=?, city=?, updated_at=datetime('now','localtime') WHERE phone=?",
                ("awaiting_quantity", custom_area, phone)
            )
        send_quantity_list(phone)
        return JSONResponse({"status": "ok"})

    # שלב 2 — בחירת כמות
    if step == "awaiting_quantity":
        if reply_id not in QUANTITY_MAP:
            send_whatsapp_message(phone, "אנא בחר כמות מהרשימה 👇")
            send_quantity_list(phone)
            return JSONResponse({"status": "ok"})
        quantity = QUANTITY_MAP[reply_id]
        with get_db() as conn:
            row = conn.execute("SELECT city FROM conversation_state WHERE phone=?", (phone,)).fetchone()
            city = row["city"] if row and row["city"] else None
            if city:
                sites = [dict(s) for s in conn.execute(
                    "SELECT * FROM customer_sites WHERE customer_id=? AND city=? AND active=1",
                    (customer["id"], city)
                ).fetchall()]
            else:
                sites = [dict(s) for s in conn.execute(
                    "SELECT * FROM customer_sites WHERE customer_id=? AND active=1",
                    (customer["id"],)
                ).fetchall()]
        if sites:
            with get_db() as conn:
                conn.execute(
                    "UPDATE conversation_state SET step=?, quantity=?, updated_at=datetime('now','localtime') WHERE phone=?",
                    ("awaiting_site", quantity, phone)
                )
            send_whatsapp_message(phone, f"מצוין! {quantity} ליטר.")
            send_site_list(phone, sites)
        else:
            with get_db() as conn:
                conn.execute(
                    "UPDATE conversation_state SET step=?, quantity=?, updated_at=datetime('now','localtime') WHERE phone=?",
                    ("awaiting_address", quantity, phone)
                )
            send_whatsapp_message(phone, f"מצוין! {quantity} ליטר.\nמה הכתובת המדויקת?")
        return JSONResponse({"status": "ok"})

    # שלב 2.5 — בחירת אתר (ללקוחות עם אתרים מוגדרים)
    if step == "awaiting_site":
        if reply_id == "site_other":
            with get_db() as conn:
                row = conn.execute("SELECT city FROM conversation_state WHERE phone=?", (phone,)).fetchone()
                city_name = row["city"] if row and row["city"] else ""
            with get_db() as conn:
                conn.execute(
                    "UPDATE conversation_state SET step=?, updated_at=datetime('now','localtime') WHERE phone=?",
                    ("awaiting_address", phone)
                )
            send_whatsapp_message(phone, f"מה הכתובת המדויקת?")
            return JSONResponse({"status": "ok"})
        if reply_id and reply_id.startswith("site_"):
            try:
                site_id = int(reply_id[5:])
            except ValueError:
                site_id = None
            if site_id:
                with get_db() as conn:
                    site = conn.execute("SELECT * FROM customer_sites WHERE id=? AND active=1", (site_id,)).fetchone()
                    if site:
                        conn.execute(
                            "UPDATE conversation_state SET step=?, city=?, address=?, updated_at=datetime('now','localtime') WHERE phone=?",
                            ("awaiting_time", site["city"], site["address"], phone)
                        )
                        send_time_list(phone)
                        return JSONResponse({"status": "ok"})
        # תשובה לא תקינה — שלח רשימה מחדש
        with get_db() as conn:
            row = conn.execute("SELECT city FROM conversation_state WHERE phone=?", (phone,)).fetchone()
            city = row["city"] if row and row["city"] else None
            if city:
                sites = [dict(s) for s in conn.execute(
                    "SELECT * FROM customer_sites WHERE customer_id=? AND city=? AND active=1",
                    (customer["id"], city)
                ).fetchall()]
            else:
                sites = [dict(s) for s in conn.execute(
                    "SELECT * FROM customer_sites WHERE customer_id=? AND active=1",
                    (customer["id"],)
                ).fetchall()]
        send_site_list(phone, sites)
        return JSONResponse({"status": "ok"})

    # שלב 3 — כתובת מדויקת
    if step == "awaiting_address":
        with get_db() as conn:
            conn.execute(
                "UPDATE conversation_state SET step=?, address=?, updated_at=datetime('now','localtime') WHERE phone=?",
                ("awaiting_time", text, phone)
            )
        send_time_list(phone)
        return JSONResponse({"status": "ok"})

    # שלב 4 — בחירת שעה → שאלת איש קשר
    if step == "awaiting_time":
        delivery_time = TIME_MAP.get(reply_id, text) if reply_id else text
        with get_db() as conn:
            conn.execute(
                "UPDATE conversation_state SET step='awaiting_contact', delivery_time=?, updated_at=datetime('now','localtime') WHERE phone=?",
                (delivery_time, phone)
            )
        send_contact_buttons(phone)
        return JSONResponse({"status": "ok"})

    # שלב 5 — בחירת איש קשר
    if step == "awaiting_contact":
        from datetime import date as dt, timedelta

        def _create_order(cname, cphone):
            with get_db() as conn:
                row        = conn.execute("SELECT * FROM conversation_state WHERE phone=?", (phone,)).fetchone()
                quantity   = row["quantity"]
                city       = row["city"]
                address    = row["address"]
                order_date = row["order_date"] or (dt.today() + timedelta(days=1)).isoformat()
                dtime      = row["delivery_time"] or ""
                driver     = assign_driver(conn, city, quantity)
                driver_id  = driver["id"] if driver else None
                conn.execute(
                    """INSERT INTO orders
                       (customer_id, customer_name, site_address, contact_name, contact_phone,
                        quantity, driver_id, order_date, delivery_time)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (customer["id"], customer["name"], f"{city}, {address}",
                     cname, cphone, quantity, driver_id, order_date, dtime)
                )
                conn.execute("DELETE FROM conversation_state WHERE phone=?", (phone,))
            _notify_admin_new_order({
                "customer_name": customer["name"], "site_address": f"{city}, {address}",
                "quantity": quantity, "order_date": order_date,
                "driver_name": driver["name"] if driver else "",
            })
            from datetime import datetime
            d = datetime.strptime(order_date, "%Y-%m-%d")
            date_label = f"{d.day}/{d.month}/{d.year}"
            send_whatsapp_message(
                phone,
                f"ההזמנה התקבלה! ✅\n\n"
                f"תאריך אספקה: {date_label}\n"
                f"כמות: {quantity} ליטר\n"
                f"כתובת: {city}, {address}\n"
                f"איש קשר: {cname}\n"
                f"טלפון נייד: {cphone}\n"
                f"שעת אספקה: {dtime}\n\n"
                f"נשתדל לעמוד בטווחי הזמנים.\n"
                f"נדאג לאספקה. תודה! 🙏\n\n"
                f"לכל בעיה נא לפנות לנאור - 0506877866"
            )

        if reply_id == "contact_self":
            _create_order(customer["name"], customer["phone"])
        elif reply_id == "contact_other":
            with get_db() as conn:
                conn.execute(
                    "UPDATE conversation_state SET step='awaiting_contact_name', updated_at=datetime('now','localtime') WHERE phone=?",
                    (phone,)
                )
            send_whatsapp_message(phone, "מה שם איש הקשר? ✍️")
        else:
            send_contact_buttons(phone)
        return JSONResponse({"status": "ok"})

    # שלב 5א — קבלת שם איש קשר
    if step == "awaiting_contact_name":
        contact_name = text.strip()
        with get_db() as conn:
            conn.execute(
                "UPDATE conversation_state SET step='awaiting_contact_phone', contact_name=?, updated_at=datetime('now','localtime') WHERE phone=?",
                (contact_name, phone)
            )
        send_whatsapp_message(phone, "מה מספר הטלפון הנייד של איש הקשר? 📱")
        return JSONResponse({"status": "ok"})

    # שלב 5ב — קבלת טלפון איש קשר → יצירת הזמנה
    if step == "awaiting_contact_phone":
        from datetime import date as dt, timedelta
        contact_phone = text.strip()
        with get_db() as conn:
            row        = conn.execute("SELECT * FROM conversation_state WHERE phone=?", (phone,)).fetchone()
            quantity   = row["quantity"]
            city       = row["city"]
            address    = row["address"]
            order_date = row["order_date"] or (dt.today() + timedelta(days=1)).isoformat()
            dtime      = row["delivery_time"] or ""
            cname      = row["contact_name"] or ""
            driver     = assign_driver(conn, city, quantity)
            driver_id  = driver["id"] if driver else None
            conn.execute(
                """INSERT INTO orders
                   (customer_id, customer_name, site_address, contact_name, contact_phone,
                    quantity, driver_id, order_date, delivery_time)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (customer["id"], customer["name"], f"{city}, {address}",
                 cname, contact_phone, quantity, driver_id, order_date, dtime)
            )
            conn.execute("DELETE FROM conversation_state WHERE phone=?", (phone,))
        _notify_admin_new_order({
            "customer_name": customer["name"], "site_address": f"{city}, {address}",
            "quantity": quantity, "order_date": order_date,
            "driver_name": driver["name"] if driver else "",
        })
        from datetime import datetime
        d = datetime.strptime(order_date, "%Y-%m-%d")
        date_label = f"{d.day}/{d.month}/{d.year}"
        send_whatsapp_message(
            phone,
            f"ההזמנה התקבלה! ✅\n\n"
            f"תאריך אספקה: {date_label}\n"
            f"כמות: {quantity} ליטר\n"
            f"כתובת: {city}, {address}\n"
            f"איש קשר: {cname}\n"
            f"טלפון נייד: {contact_phone}\n"
            f"שעת אספקה: {dtime}\n\n"
            f"נשתדל לעמוד בטווחי הזמנים.\n"
            f"נדאג לאספקה. תודה! 🙏\n\n"
            f"לכל בעיה נא לפנות לנאור - 0506877866"
        )
        return JSONResponse({"status": "ok"})

    # ברירת מחדל — הודעה לא מוכרת
    if step is None:
        greeting = time_greeting()
        send_whatsapp_message(
            phone,
            f"{greeting} {customer['name']} 😊\n"
            f"זהו הבוט של וזאת הברכה דלקים.\n"
            f"האם תרצה להזמין דלק סולר למחר?\n"
            f"ענה *כן* או *לא*"
        )
    return JSONResponse({"status": "ok"})


# ── Order Form ────────────────────────────────────────────────────────────────

@app.get("/order-form/{customer_id}", response_class=HTMLResponse)
def order_form(customer_id: int):
    return FileResponse("static/order_form.html")


@app.post("/submit-order")
def submit_order(
    customer_id: int = Form(...),
    customer_name: str = Form(...),
    site_address: str = Form(...),
    contact_name: str = Form(...),
    contact_phone: str = Form(...),
    quantity: str = Form(...),
):
    with get_db() as conn:
        customer = conn.execute(
            "SELECT * FROM customers WHERE id = ?", (customer_id,)
        ).fetchone()
        if not customer:
            raise HTTPException(status_code=404, detail="לקוח לא נמצא")

        driver = conn.execute(
            "SELECT * FROM drivers WHERE area = ? LIMIT 1", (customer["area"],)
        ).fetchone()
        driver_id = driver["id"] if driver else None

        conn.execute(
            """INSERT INTO orders
               (customer_id, customer_name, site_address, contact_name, contact_phone, quantity, driver_id)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (customer_id, customer_name, site_address, contact_name, contact_phone, quantity, driver_id),
        )
    _notify_admin_new_order({
        "customer_name": customer_name, "site_address": site_address,
        "quantity": quantity,
        "driver_name": driver["name"] if driver else "",
    })

    return HTMLResponse("""
    <!DOCTYPE html><html dir="rtl" lang="he">
    <head><meta charset="UTF-8"><title>תודה</title>
    <style>body{font-family:Arial;text-align:center;padding:60px;background:#f0f9f0;}
    h1{color:#2e7d32;}p{font-size:1.2em;}</style></head>
    <body><h1>ההזמנה התקבלה!</h1>
    <p>תודה, ההזמנה שלך נקלטה במערכת ותטופל בהקדם.</p></body></html>
    """)


# ── Admin API ─────────────────────────────────────────────────────────────────

NO_CACHE = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"}

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard():
    return FileResponse("static/dashboard.html", headers=NO_CACHE)


@app.get("/api/orders")
def get_orders(date: str = Query(default=None)):
    with get_db() as conn:
        if date:
            rows = conn.execute("""
                SELECT o.*, d.name as driver_name
                FROM orders o LEFT JOIN drivers d ON o.driver_id = d.id
                WHERE o.order_date = ?
                ORDER BY d.name, o.sort_order, o.delivery_time, o.created_at
            """, (date,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT o.*, d.name as driver_name
                FROM orders o LEFT JOIN drivers d ON o.driver_id = d.id
                ORDER BY o.created_at DESC
            """).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/orders/dates")
def get_order_dates():
    with get_db() as conn:
        rows = conn.execute("SELECT DISTINCT order_date FROM orders WHERE order_date IS NOT NULL").fetchall()
    return [r["order_date"] for r in rows]


@app.post("/api/orders")
async def add_order_manual(request: Request):
    body = await request.json()
    send_now = body.get("send_now", False)
    with get_db() as conn:
        customer_id = body.get("customer_id")
        area = body.get("area", "")
        if not area and customer_id:
            c = conn.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
            if c:
                area = c["area"]
        driver = conn.execute("SELECT * FROM drivers WHERE area = ? LIMIT 1", (area,)).fetchone()
        driver_id = driver["id"] if driver else None
        from datetime import date as dt
        order_date = body.get("order_date") or dt.today().isoformat()
        extras = body.get("extras", "")
        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) FROM orders WHERE driver_id IS ? AND order_date = ?",
            (driver_id, order_date)
        ).fetchone()[0]
        new_sort = max_sort + 1
        conn.execute(
            """INSERT INTO orders
               (customer_id, customer_name, site_address, contact_name, contact_phone, quantity, driver_id, order_date, extras, sort_order)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (customer_id, body["customer_name"], body["site_address"],
             body["contact_name"], body["contact_phone"], body["quantity"], driver_id, order_date, extras, new_sort),
        )
    _notify_admin_new_order({
        "customer_name": body["customer_name"], "site_address": body["site_address"],
        "quantity": body["quantity"], "order_date": order_date,
        "driver_name": driver["name"] if driver else "",
    })
    result = {"ok": True}
    if send_now:
        if driver:
            sent = _send_order_immediate(driver, {
                "customer_name": body["customer_name"], "site_address": body["site_address"],
                "quantity": body["quantity"], "order_date": order_date,
                "contact_name": body.get("contact_name", ""),
                "contact_phone": body.get("contact_phone", ""),
            })
            result["sent"] = sent
        else:
            result["sent"] = {"driver_ok": False, "admin_ok": False, "driver_name": "", "driver_phone": ""}
    return result


@app.get("/api/orders/latest")
def get_latest_order():
    with get_db() as conn:
        row = conn.execute(
            """SELECT o.id, o.customer_name, o.site_address, o.quantity, o.order_date, d.name as driver_name
               FROM orders o LEFT JOIN drivers d ON o.driver_id = d.id
               ORDER BY o.id DESC LIMIT 1"""
        ).fetchone()
    return dict(row) if row else {}


@app.post("/api/orders/reorder")
async def reorder_orders(request: Request):
    body = await request.json()
    ids = body.get("ids", [])
    with get_db() as conn:
        for i, order_id in enumerate(ids):
            conn.execute("UPDATE orders SET sort_order = ? WHERE id = ?", (i, order_id))
    return {"ok": True}


@app.delete("/api/orders/{order_id}")
def delete_order(order_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM orders WHERE id = ?", (order_id,))
    return {"ok": True}


@app.get("/api/recurring-orders")
def get_recurring_orders():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM recurring_orders ORDER BY created_at DESC").fetchall()
    return [dict(r) for r in rows]


@app.post("/api/recurring-orders")
async def add_recurring_order(request: Request):
    from datetime import date as _dr, timedelta as _tdr
    from recurring import materialize_recurring_orders as _mat
    body = await request.json()
    days_of_week = body.get("days_of_week") or [0, 1, 2, 3, 4]
    days_str = ",".join(str(int(d)) for d in days_of_week)
    with get_db() as conn:
        customer_id = body.get("customer_id")
        area = body.get("area", "")
        if not area and customer_id:
            c = conn.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
            if c:
                area = c["area"]
        conn.execute(
            """INSERT INTO recurring_orders
               (customer_id, customer_name, site_address, contact_name, contact_phone,
                quantity, area, days_of_week, start_date, end_date, active)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
            (customer_id, body["customer_name"], body["site_address"],
             body.get("contact_name", ""), body.get("contact_phone", ""), body["quantity"],
             area, days_str, body.get("start_date", ""), body.get("end_date", "")),
        )
    # מיד ממש הזמנות למחר (ואם צריך גם להיום) כדי שיופיעו במשימות
    for _offset in (0, 1):
        _d = (_dr.today() + _tdr(days=_offset)).isoformat()
        _mat(_d)
    return {"ok": True}


@app.put("/api/recurring-orders/{rid}")
async def update_recurring_order(rid: int, request: Request):
    body = await request.json()
    with get_db() as conn:
        if "active" in body:
            conn.execute("UPDATE recurring_orders SET active = ? WHERE id = ?", (1 if body["active"] else 0, rid))
        if "days_of_week" in body:
            days_str = ",".join(str(int(d)) for d in body["days_of_week"])
            conn.execute("UPDATE recurring_orders SET days_of_week = ? WHERE id = ?", (days_str, rid))
        if "start_date" in body:
            conn.execute("UPDATE recurring_orders SET start_date = ? WHERE id = ?", (body["start_date"], rid))
        if "end_date" in body:
            conn.execute("UPDATE recurring_orders SET end_date = ? WHERE id = ?", (body["end_date"], rid))
    return {"ok": True}


@app.delete("/api/recurring-orders/{rid}")
def delete_recurring_order(rid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM recurring_orders WHERE id = ?", (rid,))
    return {"ok": True}


@app.post("/api/recurring-orders/materialize")
async def materialize_recurring_orders_endpoint(request: Request):
    from datetime import date as dt, timedelta
    from recurring import materialize_recurring_orders
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    if body.get("target_date"):
        created = materialize_recurring_orders(body["target_date"])
    else:
        created = 0
        for offset in (0, 1, 2):
            created += materialize_recurring_orders((dt.today() + timedelta(days=offset)).isoformat())
    return {"ok": True, "created": created}


@app.put("/api/orders/{order_id}/status")
async def update_order_status(order_id: int, request: Request):
    body = await request.json()
    status = body.get("status")
    with get_db() as conn:
        conn.execute("UPDATE orders SET status = ? WHERE id = ?", (status, order_id))
    return {"ok": True}


@app.put("/api/orders/{order_id}")
async def update_order(order_id: int, request: Request):
    body = await request.json()
    with get_db() as conn:
        if "driver_id" in body:
            driver_id = body["driver_id"]  # ישירות מהטופס
        else:
            area = body.get("area", "")
            driver = conn.execute("SELECT id FROM drivers WHERE area = ? LIMIT 1", (area,)).fetchone()
            driver_id = driver["id"] if driver else None
        conn.execute(
            """UPDATE orders SET
               customer_name = ?, site_address = ?, contact_name = ?, contact_phone = ?,
               quantity = ?, order_date = ?, driver_id = ?, extras = ?
               WHERE id = ?""",
            (body.get("customer_name"), body.get("site_address"),
             body.get("contact_name", ""), body.get("contact_phone", ""),
             body.get("quantity"), body.get("order_date"),
             driver_id, body.get("extras", ""), order_id),
        )
    return {"ok": True}


# ── רישומים ממתינים ────────────────────────────────────────────────────────

@app.get("/api/pending-registrations")
def get_pending_registrations():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM pending_registrations WHERE status='pending' ORDER BY created_at DESC"
        ).fetchall()
    return [dict(r) for r in rows]

@app.post("/api/pending-registrations/{pid}/approve")
async def approve_pending(pid: int, request: Request):
    body = await request.json()
    with get_db() as conn:
        pend = conn.execute("SELECT * FROM pending_registrations WHERE id=?", (pid,)).fetchone()
        if not pend:
            raise HTTPException(status_code=404, detail="לא נמצא")
        conn.execute(
            "INSERT INTO customers (name, phone, area, site_address, contact_name, contact_phone) VALUES (?,?,?,?,?,?)",
            (body.get("name", pend["company_name"]), pend["phone"],
             body.get("area", ""), body.get("site_address", ""),
             body.get("contact_name", ""), body.get("contact_phone", ""))
        )
        conn.execute("UPDATE pending_registrations SET status='approved' WHERE id=?", (pid,))
    send_whatsapp_message(
        pend["phone"],
        f"שלום {body.get('name', pend['company_name'])}! 👋\n"
        "נרשמת בהצלחה במערכת וזאת הברכה דלקים.\n"
        "כעת תוכל להזמין דלק — פשוט שלח *הזמנה* 😊"
    )
    return {"ok": True}

@app.post("/api/pending-registrations/{pid}/link")
async def link_pending(pid: int, request: Request):
    body = await request.json()
    customer_id = body.get("customer_id")
    if not customer_id:
        raise HTTPException(status_code=400, detail="חסר customer_id")
    with get_db() as conn:
        pend = conn.execute("SELECT * FROM pending_registrations WHERE id=?", (pid,)).fetchone()
        if not pend:
            raise HTTPException(status_code=404, detail="לא נמצא")
        customer = conn.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
        if not customer:
            raise HTTPException(status_code=404, detail="לקוח לא נמצא")
        contact_name = body.get("contact_name", "")
        # הוסף את הטלפון החדש — אם שדה הטלפון הקיים ריק, מלא אותו; אחרת שמור כ-phone2
        if not customer["phone"]:
            conn.execute("UPDATE customers SET phone=? WHERE id=?", (pend["phone"], customer_id))
        else:
            conn.execute("UPDATE customers SET phone2=? WHERE id=?", (pend["phone"], customer_id))
        # עדכן איש קשר אם הוזן
        if contact_name:
            conn.execute("UPDATE customers SET contact_name=? WHERE id=?", (contact_name, customer_id))
        conn.execute("UPDATE pending_registrations SET status='approved' WHERE id=?", (pid,))
    greeting_name = contact_name or customer['name']
    send_whatsapp_message(
        pend["phone"],
        f"שלום {greeting_name}! 👋\n"
        "מספרך שויך בהצלחה במערכת וזאת הברכה דלקים.\n"
        "כעת תוכל להזמין דלק — פשוט שלח *הזמנה* 😊"
    )
    return {"ok": True}

@app.delete("/api/pending-registrations/{pid}")
def delete_pending(pid: int):
    with get_db() as conn:
        conn.execute("UPDATE pending_registrations SET status='rejected' WHERE id=?", (pid,))
    return {"ok": True}


@app.get("/api/customers")
def get_customers():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT c.*,
                   COALESCE(lo.contact_name, '') as last_order_contact_name,
                   COALESCE(lo.contact_phone, '') as last_order_contact_phone
            FROM customers c
            LEFT JOIN orders lo ON lo.id = (
                SELECT id FROM orders
                WHERE customer_id = c.id
                  AND (contact_name != '' OR contact_phone != '')
                ORDER BY order_date DESC, id DESC
                LIMIT 1
            )
            WHERE c.active=1
            ORDER BY c.name
        """).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/customers")
async def add_customer(request: Request):
    body = await request.json()
    with get_db() as conn:
        conn.execute(
            "INSERT INTO customers (name, phone, area, site_address, contact_name, contact_phone, email, order_contact_name, order_contact_phone) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (body["name"], body.get("phone", ""), body.get("area", ""),
             body.get("site_address", ""), body.get("contact_name", ""), body.get("contact_phone", ""),
             body.get("email", ""), body.get("order_contact_name", ""), body.get("order_contact_phone", "")),
        )
    return {"ok": True}


@app.put("/api/customers/{cid}")
async def update_customer(cid: int, request: Request):
    body = await request.json()
    with get_db() as conn:
        conn.execute(
            "UPDATE customers SET name=?, phone=?, area=?, site_address=?, contact_name=?, contact_phone=?, email=?, order_contact_name=?, order_contact_phone=? WHERE id=?",
            (body["name"], body.get("phone", ""), body.get("area", ""),
             body.get("site_address", ""), body.get("contact_name", ""),
             body.get("contact_phone", ""), body.get("email", ""),
             body.get("order_contact_name", ""), body.get("order_contact_phone", ""), cid),
        )
    return {"ok": True}


@app.delete("/api/customers/{cid}")
def delete_customer(cid: int):
    with get_db() as conn:
        conn.execute("UPDATE customers SET active = 0 WHERE id = ?", (cid,))
    return {"ok": True}


@app.post("/api/customers/{cid}/send-order-message")
def send_order_message(cid: int):
    with get_db() as conn:
        customer = conn.execute("SELECT * FROM customers WHERE id=? AND active=1", (cid,)).fetchone()
        if not customer:
            raise HTTPException(status_code=404, detail="לקוח לא נמצא")
        settings = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
    template = settings.get("message_template", "שלום {name}, האם תצטרך הזמנת דלק סולר למחר?")
    from whatsapp import send_daily_question
    phone = customer["phone"].strip().replace(" ", "").replace("-", "")
    if phone.startswith("0"):
        phone = "972" + phone[1:]
    send_daily_question(phone, customer["name"], template)
    return {"ok": True}


@app.get("/api/customers/{cid}/sites")
def get_customer_sites(cid: int):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM customer_sites WHERE customer_id=? AND active=1 ORDER BY id",
            (cid,)
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/customers/{cid}/sites")
async def add_customer_site(cid: int, request: Request):
    body = await request.json()
    name    = body.get("name", "").strip()
    city    = body.get("city", "").strip()
    address = body.get("address", "").strip()
    if not name or not city or not address:
        raise HTTPException(status_code=400, detail="name, city, address חובה")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO customer_sites (customer_id, name, city, address) VALUES (?,?,?,?)",
            (cid, name, city, address)
        )
    return {"ok": True}


@app.delete("/api/customers/{cid}/sites/{sid}")
def delete_customer_site(cid: int, sid: int):
    with get_db() as conn:
        conn.execute("UPDATE customer_sites SET active=0 WHERE id=? AND customer_id=?", (sid, cid))
    return {"ok": True}


@app.get("/api/drivers")
def get_drivers():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM drivers ORDER BY name").fetchall()
    return [dict(r) for r in rows]


@app.post("/api/drivers")
async def add_driver(request: Request):
    body = await request.json()
    with get_db() as conn:
        conn.execute(
            "INSERT INTO drivers (name, phone, area, tanker_volume, truck_number, personal_phone) VALUES (?, ?, ?, ?, ?, ?)",
            (body["name"], body["phone"], body["area"],
             body.get("tanker_volume", ""), body.get("truck_number", ""),
             body.get("personal_phone", "")),
        )
    return {"ok": True}


@app.get("/api/drivers/{did}/orders")
def get_driver_orders(did: int, date: str = Query(default=None)):
    from datetime import date as dt
    target = date or dt.today().isoformat()
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM orders
            WHERE driver_id = ? AND order_date = ?
            ORDER BY created_at
        """, (did, target)).fetchall()
    return [dict(r) for r in rows]


@app.put("/api/drivers/{did}")
async def update_driver(did: int, request: Request):
    body = await request.json()
    with get_db() as conn:
        conn.execute(
            "UPDATE drivers SET name=?, phone=?, area=?, truck_number=?, tanker_volume=?, personal_phone=? WHERE id=?",
            (body["name"], body["phone"], body["area"],
             body.get("truck_number", ""), body.get("tanker_volume", ""),
             body.get("personal_phone", ""), did),
        )
    return {"ok": True}


@app.delete("/api/drivers/{did}")
def delete_driver(did: int):
    with get_db() as conn:
        conn.execute("DELETE FROM drivers WHERE id = ?", (did,))
    return {"ok": True}


@app.get("/api/settings")
def get_settings():
    with get_db() as conn:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
    return {r["key"]: r["value"] for r in rows}


@app.post("/api/settings")
async def update_settings(request: Request):
    body = await request.json()
    with get_db() as conn:
        for key, value in body.items():
            conn.execute(
                "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
                (key, str(value)),
            )
    if "daily_hour" in body or "daily_minute" in body:
        with get_db() as conn:
            s = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
        reschedule(int(s["daily_hour"]), int(s["daily_minute"]))
    if "admin_schedule_hour" in body or "admin_schedule_minute" in body:
        with get_db() as conn:
            s = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
        reschedule_admin(int(s["admin_schedule_hour"]), int(s.get("admin_schedule_minute", 0)))
    return {"ok": True}


@app.get("/api/tasks/week")
def get_weekly_tasks():
    from datetime import date as dt, timedelta
    today = dt.today()
    week_dates = [(today + timedelta(days=i)).isoformat() for i in range(7)]
    with get_db() as conn:
        drivers = conn.execute("SELECT * FROM drivers ORDER BY name").fetchall()
        result = []
        for d in drivers:
            orders_by_date = {}
            for date_str in week_dates:
                rows = conn.execute("""
                    SELECT * FROM orders WHERE driver_id = ? AND order_date = ?
                    ORDER BY sort_order, delivery_time, created_at
                """, (d["id"], date_str)).fetchall()
                if rows:
                    seen = set()
                    unique = []
                    for r in rows:
                        key = (r["customer_name"], r["site_address"])
                        if key not in seen:
                            seen.add(key)
                            unique.append(dict(r))
                    orders_by_date[date_str] = unique
            result.append({**dict(d), "orders_by_date": orders_by_date})

        # הזמנות ללא נהג משויך
        unassigned_by_date = {}
        for date_str in week_dates:
            rows = conn.execute("""
                SELECT * FROM orders WHERE driver_id IS NULL AND order_date = ?
                ORDER BY created_at
            """, (date_str,)).fetchall()
            if rows:
                seen = set()
                unique = []
                for r in rows:
                    key = (r["customer_name"], r["site_address"])
                    if key not in seen:
                        seen.add(key)
                        unique.append(dict(r))
                unassigned_by_date[date_str] = unique
        if unassigned_by_date:
            result.append({
                "id": None, "name": "ללא נהג", "area": "", "phone": "",
                "personal_phone": "", "truck_number": "", "tanker_volume": "",
                "orders_by_date": unassigned_by_date,
            })

    return result


# ─── Terminal Trips ────────────────────────────────────────────────────────────

@app.get("/api/terminal-trips")
def list_terminal_trips(date: str = None):
    with get_db() as conn:
        if date:
            rows = conn.execute("""
                SELECT t.*, d.name AS driver_name
                FROM terminal_trips t
                LEFT JOIN drivers d ON t.driver_id = d.id
                WHERE t.trip_date = ?
                ORDER BY t.created_at DESC
            """, (date,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT t.*, d.name AS driver_name
                FROM terminal_trips t
                LEFT JOIN drivers d ON t.driver_id = d.id
                ORDER BY t.trip_date DESC, t.created_at DESC
                LIMIT 60
            """).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/terminal-trips")
async def create_terminal_trip(request: Request):
    import json as _json
    data = await request.json()
    driver_id    = data.get("driver_id")
    trip_date    = data.get("trip_date", "")
    fuel_company = data.get("fuel_company", "")
    cert_number  = data.get("cert_number", "")
    compartments = _json.dumps(data.get("compartments", []), ensure_ascii=False)
    notes        = data.get("notes", "")
    send_wa      = data.get("send_whatsapp", False)

    with get_db() as conn:
        conn.execute(
            "INSERT INTO terminal_trips (driver_id, trip_date, fuel_company, cert_number, compartments, notes) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (driver_id, trip_date, fuel_company, cert_number, compartments, notes)
        )
        trip_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        wa_sent = False
        wa_error = ""
        if send_wa and driver_id:
            driver = conn.execute("SELECT * FROM drivers WHERE id = ?", (driver_id,)).fetchone()
            if driver:
                driver = dict(driver)
                phone = _fmt_phone(driver.get("personal_phone") or driver.get("phone") or "")
                if phone:
                    comps = data.get("compartments", [])
                    total = sum(float(c.get("qty") or 0) for c in comps)
                    lines = [f"⛽ *נסיעה למסוף — {trip_date}*",
                             f"חברת דלק: *{fuel_company}*"]
                    if cert_number:
                        lines.append(f"מס' תעודה: *{cert_number}*")
                    lines.append("\n*תאים:*")
                    for c in comps:
                        lines.append(f"  תא {c.get('num','')} ← {c.get('qty','')} ל'")
                    lines.append(f"\n*סה\"כ: {int(total):,} ל'*")
                    if notes:
                        lines.append(f"\nהערות: {notes}")
                    try:
                        send_whatsapp_message(phone, "\n".join(lines))
                        wa_sent = True
                    except Exception as e:
                        wa_error = str(e)

    return {"ok": True, "id": trip_id, "wa_sent": wa_sent, "wa_error": wa_error}


@app.put("/api/terminal-trips/{trip_id}")
async def update_terminal_trip(trip_id: int, request: Request):
    import json as _json
    data = await request.json()
    trip_date    = data.get("trip_date", "")
    fuel_company = data.get("fuel_company", "")
    cert_number  = data.get("cert_number", "")
    compartments = _json.dumps(data.get("compartments", []), ensure_ascii=False)
    notes        = data.get("notes", "")
    send_wa      = data.get("send_whatsapp", False)

    with get_db() as conn:
        conn.execute("""
            UPDATE terminal_trips
            SET trip_date=?, fuel_company=?, cert_number=?, compartments=?, notes=?
            WHERE id=?
        """, (trip_date, fuel_company, cert_number, compartments, notes, trip_id))

        wa_sent = False
        wa_error = ""
        if send_wa:
            row = conn.execute("SELECT t.*, d.personal_phone, d.phone FROM terminal_trips t LEFT JOIN drivers d ON t.driver_id=d.id WHERE t.id=?", (trip_id,)).fetchone()
            if row:
                row = dict(row)
                phone = _fmt_phone(row.get("personal_phone") or row.get("phone") or "")
                if phone:
                    comps = data.get("compartments", [])
                    total = sum(float(c.get("qty") or 0) for c in comps)
                    lines = [f"⛽ *נסיעה למסוף — {trip_date}*", f"חברת דלק: *{fuel_company}*"]
                    if cert_number:
                        lines.append(f"מס' תעודה: *{cert_number}*")
                    lines.append("\n*תאים:*")
                    for c in comps:
                        lines.append(f"  תא {c.get('num','')} ← {c.get('qty','')} ל'")
                    lines.append(f"\n*סה\"כ: {int(total):,} ל'*")
                    try:
                        send_whatsapp_message(phone, "\n".join(lines))
                        wa_sent = True
                    except Exception as e:
                        wa_error = str(e)

    return {"ok": True, "wa_sent": wa_sent, "wa_error": wa_error}


@app.delete("/api/terminal-trips/{trip_id}")
def delete_terminal_trip(trip_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM terminal_trips WHERE id = ?", (trip_id,))
    return {"ok": True}


@app.patch("/api/drivers/{driver_id}/terminal")
async def toggle_terminal_driver(driver_id: int, request: Request):
    data = await request.json()
    val = 1 if data.get("is_terminal_driver") else 0
    with get_db() as conn:
        conn.execute("UPDATE drivers SET is_terminal_driver = ? WHERE id = ?", (val, driver_id))
    return {"ok": True}


@app.post("/api/orders/dedup")
def dedup_orders():
    """מחיקת הזמנות כפולות — שומר רק את זו עם ה-id הנמוך ביותר (ללא תלות בנהג)"""
    with get_db() as conn:
        deleted = conn.execute("""
            DELETE FROM orders
            WHERE id NOT IN (
                SELECT MIN(id)
                FROM orders
                GROUP BY order_date, customer_name, site_address
            )
        """).rowcount
    return {"ok": True, "deleted": deleted}


@app.post("/api/customers/import-excel")
async def import_excel(request: Request):
    import openpyxl, io
    body = await request.body()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(body))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"לא ניתן לפתוח את הקובץ: {e}")

    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    COL_MAP = {
        "name":          ["שם","שם לקוח","name","company","שם עסק","לקוח"],
        "phone":         ["טלפון","phone","mobile","נייד","טל","טל'"],
        "area":          ["אזור","area","region","אזור חלוקה"],
        "email":         ["מייל","email","e-mail","אימייל","דואר אלקטרוני"],
        "site_address":  ["כתובת","כתובת אתר","address","site","כתובת האתר"],
        "contact_name":  ["איש קשר","contact","contact name","שם איש קשר"],
        "contact_phone": ["טלפון איש קשר","contact phone","נייד איש קשר"],
    }

    col_idx = {}
    for field, aliases in COL_MAP.items():
        for i, h in enumerate(headers):
            if h.strip().lower() in [a.lower() for a in aliases]:
                col_idx[field] = i
                break

    if "name" not in col_idx:
        raise HTTPException(
            status_code=400,
            detail=f"לא נמצאה עמודת שם לקוח. כותרות שנמצאו בקובץ: {', '.join(h for h in headers if h)}"
        )

    imported = 0
    skipped = 0
    with get_db() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            def g(field):
                idx = col_idx.get(field)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else ""
            name = g("name")
            if not name or name == "None":
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO customers (name, phone, area, email, site_address, contact_name, contact_phone) VALUES (?,?,?,?,?,?,?)",
                (name, g("phone"), g("area"), g("email"), g("site_address"), g("contact_name"), g("contact_phone"))
            )
            imported += 1

    return {"ok": True, "imported": imported, "skipped": skipped,
            "mapped_columns": list(col_idx.keys())}


@app.get("/api/potential-customers")
def get_potential_customers(q: str = Query(default=""), limit: int = Query(default=50), offset: int = Query(default=0)):
    with get_db() as conn:
        order = "ORDER BY (notes IS NOT NULL AND notes != '') DESC, name"
        if q:
            pattern = f"%{q}%"
            total = conn.execute(
                "SELECT COUNT(*) FROM potential_customers WHERE active=1 AND (name LIKE ? OR phone LIKE ? OR area LIKE ?)",
                (pattern, pattern, pattern)
            ).fetchone()[0]
            rows = conn.execute(
                f"SELECT * FROM potential_customers WHERE active=1 AND (name LIKE ? OR phone LIKE ? OR area LIKE ?) {order} LIMIT ? OFFSET ?",
                (pattern, pattern, pattern, limit, offset)
            ).fetchall()
        else:
            total = conn.execute("SELECT COUNT(*) FROM potential_customers WHERE active=1").fetchone()[0]
            rows = conn.execute(
                f"SELECT * FROM potential_customers WHERE active=1 {order} LIMIT ? OFFSET ?",
                (limit, offset)
            ).fetchall()
    return {"items": [dict(r) for r in rows], "total": total}


@app.post("/api/potential-customers")
async def add_potential_customer(request: Request):
    body = await request.json()
    with get_db() as conn:
        conn.execute(
            "INSERT INTO potential_customers (name, phone, area, site_address, contact_name, contact_phone, email) VALUES (?,?,?,?,?,?,?)",
            (body["name"], body.get("phone",""), body.get("area",""),
             body.get("site_address",""), body.get("contact_name",""),
             body.get("contact_phone",""), body.get("email","")),
        )
    return {"ok": True}


@app.put("/api/potential-customers/{cid}/notes")
async def update_potential_customer_notes(cid: int, request: Request):
    body = await request.json()
    with get_db() as conn:
        conn.execute("UPDATE potential_customers SET notes = ? WHERE id = ?", (body.get("notes", ""), cid))
    return {"ok": True}


@app.delete("/api/potential-customers/{cid}")
def delete_potential_customer(cid: int):
    with get_db() as conn:
        conn.execute("UPDATE potential_customers SET active = 0 WHERE id = ?", (cid,))
    return {"ok": True}


@app.post("/api/potential-customers/import-excel")
async def import_potential_excel(request: Request):
    import openpyxl, io
    body = await request.body()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(body))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"לא ניתן לפתוח את הקובץ: {e}")

    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    COL_MAP = {
        "name":          ["שם","שם לקוח","name","company","שם עסק","לקוח"],
        "phone":         ["טלפון","phone","mobile","נייד","טל","טל'","טלפון ראשי"],
        "area":          ["אזור","area","region","אזור חלוקה"],
        "email":         ["מייל","email","e-mail","אימייל","דואר אלקטרוני"],
        "site_address":  ["כתובת","כתובת אתר","address","site","כתובת האתר"],
        "contact_name":  ["איש קשר","contact","contact name","שם איש קשר"],
        "contact_phone": ["טלפון איש קשר","contact phone","נייד איש קשר"],
    }

    col_idx = {}
    for field, aliases in COL_MAP.items():
        for i, h in enumerate(headers):
            if h.strip().lower() in [a.lower() for a in aliases]:
                col_idx[field] = i
                break

    if "name" not in col_idx:
        raise HTTPException(
            status_code=400,
            detail=f"לא נמצאה עמודת שם לקוח. כותרות שנמצאו: {', '.join(h for h in headers if h)}"
        )

    imported = 0
    skipped = 0
    with get_db() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            def g(field):
                idx = col_idx.get(field)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else ""
            name = g("name")
            if not name or name == "None":
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO potential_customers (name, phone, area, email, site_address, contact_name, contact_phone) VALUES (?,?,?,?,?,?,?)",
                (name, g("phone"), g("area"), g("email"), g("site_address"), g("contact_name"), g("contact_phone"))
            )
            imported += 1

    return {"ok": True, "imported": imported, "skipped": skipped,
            "mapped_columns": list(col_idx.keys())}


_campaign_status: dict = {"running": False, "sent": 0, "failed": 0, "total": 0, "done": True, "errors": []}


@app.get("/api/campaign-status")
def campaign_status():
    return _campaign_status


@app.post("/api/send-campaign")
async def send_campaign(request: Request, background_tasks=None):
    import smtplib, threading
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    body = await request.json()
    subject   = body.get("subject", "")
    content   = body.get("body", "")
    test_email = body.get("test_email", "").strip()

    with get_db() as conn:
        s = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}

    GMAIL_USER = s.get("gmail_user", "") or os.getenv("GMAIL_USER", "")
    GMAIL_PASS = s.get("gmail_app_password", "") or os.getenv("GMAIL_APP_PASSWORD", "")
    APP_URL    = (s.get("app_url", "") or "").rstrip("/")

    if not GMAIL_USER or not GMAIL_PASS:
        return {"ok": False, "error": "חסרים פרטי Gmail — הגדר בלשונית הגדרות"}

    if test_email:
        recipients = [{"email": test_email, "name": "בדיקה", "customer_id": None}]
    else:
        with get_db() as conn:
            cust  = conn.execute("SELECT id, name, email FROM customers WHERE active=1 AND email != '' AND email IS NOT NULL").fetchall()
            pcust = conn.execute("SELECT id, name, email FROM potential_customers WHERE active=1 AND email != '' AND email IS NOT NULL").fetchall()
        recipients = [{"email": r["email"], "name": r["name"], "customer_id": r["id"]} for r in list(cust) + list(pcust)]

    if not recipients:
        return {"ok": False, "error": "אין לקוחות עם כתובת מייל"}

    with get_db() as conn:
        cur = conn.execute("INSERT INTO email_campaigns (subject, content) VALUES (?, ?)", (subject, content))
        campaign_id = cur.lastrowid
        for r in recipients:
            conn.execute(
                "INSERT INTO email_campaign_recipients (campaign_id, customer_id, name, email) VALUES (?,?,?,?)",
                (campaign_id, r.get("customer_id"), r.get("name", ""), r["email"])
            )

    with get_db() as conn:
        rec_rows = [dict(r) for r in conn.execute(
            "SELECT id, email, name FROM email_campaign_recipients WHERE campaign_id=?", (campaign_id,)
        ).fetchall()]

    _campaign_status.update({"running": True, "sent": 0, "failed": 0, "total": len(rec_rows), "done": False, "errors": []})

    def _send_all():
        try:
            smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30)
            smtp.login(GMAIL_USER, GMAIL_PASS)
            for rec in rec_rows:
                try:
                    msg = MIMEMultipart("alternative")
                    msg["Subject"] = subject
                    msg["From"]    = GMAIL_USER
                    msg["To"]      = rec["email"]
                    personal = content.replace("{name}", rec["name"])
                    pixel = ""
                    if APP_URL and not test_email:
                        pixel = f'<img src="{APP_URL}/api/email-track/{campaign_id}/{rec["id"]}" width="1" height="1" style="display:none">'
                    html_body = personal.replace("\n", "<br>") + pixel
                    msg.attach(MIMEText(personal, "plain", "utf-8"))
                    msg.attach(MIMEText(f"<html><body>{html_body}</body></html>", "html", "utf-8"))
                    smtp.sendmail(GMAIL_USER, rec["email"], msg.as_string())
                    _campaign_status["sent"] += 1
                except Exception as e:
                    _campaign_status["failed"] += 1
                    _campaign_status["errors"].append(rec["email"])
                    with get_db() as conn:
                        conn.execute("UPDATE email_campaign_recipients SET status='failed' WHERE id=?", (rec["id"],))
            smtp.quit()
        except Exception as e:
            _campaign_status["errors"].append(f"שגיאת SMTP: {e}")
            _campaign_status["failed"] = _campaign_status["total"] - _campaign_status["sent"]
        finally:
            with get_db() as conn:
                conn.execute("UPDATE email_campaigns SET total_sent=?, total_failed=? WHERE id=?",
                             (_campaign_status["sent"], _campaign_status["failed"], campaign_id))
            _campaign_status["running"] = False
            _campaign_status["done"]    = True

    threading.Thread(target=_send_all, daemon=True).start()
    return {"ok": True, "total": len(rec_rows), "campaign_id": campaign_id}


@app.post("/api/send-email")
async def send_mass_email(request: Request):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    body = await request.json()
    subject    = body.get("subject", "")
    content    = body.get("content", "")
    recipients = body.get("recipients", [])
    test_only  = body.get("test_email", "")  # אם מוגדר — שלח רק אליו

    with get_db() as conn:
        s = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}

    GMAIL_USER = s.get("gmail_user", "") or os.getenv("GMAIL_USER", "")
    GMAIL_PASS = s.get("gmail_app_password", "") or os.getenv("GMAIL_APP_PASSWORD", "")
    APP_URL    = (s.get("app_url", "") or "").rstrip("/")

    if not GMAIL_USER or not GMAIL_PASS:
        raise HTTPException(status_code=500, detail="חסרים פרטי Gmail — הגדר בלשונית הגדרות")

    if test_only:
        recipients = [{"email": test_only, "name": "בדיקה", "customer_id": None}]

    if not recipients:
        raise HTTPException(status_code=400, detail="אין לקוחות עם כתובת מייל")

    # צור רשומת קמפיין
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO email_campaigns (subject, content) VALUES (?, ?)",
            (subject, content)
        )
        campaign_id = cur.lastrowid
        for r in recipients:
            conn.execute(
                "INSERT INTO email_campaign_recipients (campaign_id, customer_id, name, email) VALUES (?,?,?,?)",
                (campaign_id, r.get("customer_id"), r.get("name", ""), r["email"])
            )

    # שלוף את מזהי הנמענים
    with get_db() as conn:
        rec_rows = conn.execute(
            "SELECT id, email, name FROM email_campaign_recipients WHERE campaign_id=?",
            (campaign_id,)
        ).fetchall()

    sent, failed = 0, 0
    try:
        smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30)
        smtp.login(GMAIL_USER, GMAIL_PASS)

        for rec in rec_rows:
            try:
                msg = MIMEMultipart("alternative")
                msg["Subject"] = subject
                msg["From"]    = GMAIL_USER
                msg["To"]      = rec["email"]
                personal = content.replace("{name}", rec["name"])

                # HTML עם pixel מעקב
                pixel = ""
                if APP_URL and not test_only:
                    pixel = f'<img src="{APP_URL}/api/email-track/{campaign_id}/{rec["id"]}" width="1" height="1" style="display:none">'
                html_body = personal.replace("\n", "<br>") + pixel
                msg.attach(MIMEText(personal, "plain", "utf-8"))
                msg.attach(MIMEText(f"<html><body>{html_body}</body></html>", "html", "utf-8"))

                smtp.sendmail(GMAIL_USER, rec["email"], msg.as_string())
                sent += 1
            except Exception:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE email_campaign_recipients SET status='failed' WHERE id=?",
                        (rec["id"],)
                    )
                failed += 1

        smtp.quit()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאת Gmail: {e}")

    with get_db() as conn:
        conn.execute(
            "UPDATE email_campaigns SET total_sent=?, total_failed=? WHERE id=?",
            (sent, failed, campaign_id)
        )

    return {"ok": True, "sent": sent, "failed": failed, "campaign_id": campaign_id}


@app.get("/api/email-track/{campaign_id}/{recipient_id}")
def track_email_open(campaign_id: int, recipient_id: int):
    from fastapi.responses import Response
    with get_db() as conn:
        conn.execute(
            "UPDATE email_campaign_recipients SET status='opened', opened_at=datetime('now','localtime') WHERE id=? AND status='sent'",
            (recipient_id,)
        )
        conn.execute(
            "UPDATE email_campaigns SET total_opened = total_opened + 1 WHERE id=? AND EXISTS (SELECT 1 FROM email_campaign_recipients WHERE id=? AND status='opened' AND opened_at IS NOT NULL)",
            (campaign_id, recipient_id)
        )
    # 1x1 שקוף GIF
    gif = b"GIF89a\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00!\xf9\x04\x00\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;"
    return Response(content=gif, media_type="image/gif")


@app.get("/api/email-campaigns")
def get_email_campaigns():
    with get_db() as conn:
        campaigns = conn.execute(
            "SELECT * FROM email_campaigns ORDER BY sent_at DESC LIMIT 20"
        ).fetchall()
        result = []
        for c in campaigns:
            recs = conn.execute(
                "SELECT name, email, status, opened_at FROM email_campaign_recipients WHERE campaign_id=?",
                (c["id"],)
            ).fetchall()
            result.append({**dict(c), "recipients": [dict(r) for r in recs]})
    return result


@app.get("/api/customers/{cid}/contacts")
def get_customer_contacts(cid: int):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM customer_contacts WHERE customer_id=? AND active=1 ORDER BY created_at",
            (cid,)
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/customers/{cid}/contacts")
async def add_customer_contact(cid: int, request: Request):
    body = await request.json()
    name  = body.get("name", "").strip()
    phone = body.get("phone", "").strip().replace("-", "").replace(" ", "")
    send_invite = body.get("send_invite", False)
    if not name or not phone:
        raise HTTPException(status_code=400, detail="שם וטלפון חובה")
    with get_db() as conn:
        customer = conn.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
        if not customer:
            raise HTTPException(status_code=404, detail="לקוח לא נמצא")
        conn.execute(
            "INSERT INTO customer_contacts (customer_id, name, phone) VALUES (?,?,?)",
            (cid, name, phone)
        )
    if send_invite:
        wa_phone = ("972" + phone[1:]) if phone.startswith("0") else phone
        send_whatsapp_message(
            wa_phone,
            f"שלום {name}! 👋\n"
            f"כעת אתה מורשה להזמין דלק עבור *{customer['name']}* דרך הבוט שלנו.\n\n"
            f"פשוט שלח *הזמנה* ואנחנו נטפל בשאר 😊"
        )
    return {"ok": True}


@app.delete("/api/customers/{cid}/contacts/{contact_id}")
def delete_customer_contact(cid: int, contact_id: int):
    with get_db() as conn:
        conn.execute(
            "UPDATE customer_contacts SET active=0 WHERE id=? AND customer_id=?",
            (contact_id, cid)
        )
    return {"ok": True}


@app.post("/api/test-message")
async def send_test_message(request: Request):
    body = await request.json()
    phone = body.get("phone", "").strip().replace(" ", "").replace("-", "")
    if not phone:
        raise HTTPException(status_code=400, detail="חסר מספר טלפון")
    if phone.startswith("0"):
        phone = "972" + phone[1:]
    with get_db() as conn:
        settings = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
    template = settings.get("message_template", "שלום {name}, האם תצטרך הזמנת דלק סולר למחר?")
    text = template.replace("{name}", "לקוח לדוגמה")
    from whatsapp import send_whatsapp_message
    success = send_whatsapp_message(phone, text)
    return {"ok": success, "message": text}


@app.get("/api/debug-schedule")
def debug_schedule():
    """מידע אבחון: מספר מנהל + כמות הזמנות למחר (ללא שליחה)."""
    from datetime import date as dt, timedelta
    target_date = (dt.today() + timedelta(days=1)).isoformat()
    with get_db() as conn:
        settings    = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
        admin_phone_raw = settings.get("admin_phone", "")
        count = conn.execute("SELECT COUNT(*) FROM orders WHERE order_date=?", (target_date,)).fetchone()[0]
    phone = admin_phone_raw.strip().replace(" ", "").replace("-", "")
    if phone.startswith("0"):
        phone = "972" + phone[1:]
    return {
        "admin_phone_stored": admin_phone_raw,
        "admin_phone_formatted": phone,
        "target_date": target_date,
        "orders_count": count,
        "phone_ok": bool(phone),
    }


@app.post("/api/send-daily-schedule")
async def send_daily_schedule(request: Request):
    from datetime import date as dt, timedelta
    from whatsapp import send_order_card
    target_date = (dt.today() + timedelta(days=1)).isoformat()

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    # ordered_ids: [[id1, id2, ...], [id3, id4, ...], ...] — מהלקוח לפי נהג
    ordered_ids: list = body.get("ordered_ids", [])

    with get_db() as conn:
        settings    = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM settings").fetchall()}
        admin_phone = settings.get("admin_phone", "").strip().replace(" ", "").replace("-", "")
        if admin_phone.startswith("0"):
            admin_phone = "972" + admin_phone[1:]
        if not admin_phone:
            return {"ok": False, "error": "לא הוגדר טלפון מנהל"}

        orders_raw = conn.execute("""
            SELECT o.*, d.name as driver_name, d.phone as driver_phone,
                   d.personal_phone as driver_personal_phone
            FROM orders o LEFT JOIN drivers d ON o.driver_id = d.id
            WHERE o.order_date = ?
            ORDER BY d.name, o.sort_order, o.delivery_time, o.created_at
        """, (target_date,)).fetchall()
        orders_by_id = {o["id"]: dict(o) for o in orders_raw}

    if not orders_by_id:
        wa_ok = send_whatsapp_message(admin_phone, f"📋 סידור יומי — {target_date}\n\nאין הזמנות למחר.")
        return {"ok": wa_ok, "orders_sent": 0,
                "error": None if wa_ok else f"וואטסאפ נכשל — בדוק token ומספר מנהל ({admin_phone})"}

    # אם הלקוח שלח סדר מפורש — השתמש בו; אחרת — sort_order מה-DB
    if ordered_ids:
        flat_ids = [oid for group in ordered_ids for oid in group]
        orders = [orders_by_id[oid] for oid in flat_ids if oid in orders_by_id]
        # הוסף הזמנות שלא נכללו בסדר (אם יש)
        included = set(flat_ids)
        orders += [o for o in orders_by_id.values() if o["id"] not in included]
    else:
        orders = sorted(orders_by_id.values(),
                        key=lambda o: (o.get("driver_name") or "", o.get("sort_order") or 0))

    # קבץ לפי נהג — תוך שמירת הסדר
    by_driver = {}
    for o in orders:
        name = o["driver_name"] or "ללא נהג"
        by_driver.setdefault(name, []).append(o)

    # שלח למנהל — רשימת טקסט ממוספרת (ללא כרטיסים)
    admin_lines = [f"📋 *{target_date}* — {len(orders)} הזמנות\n"]
    for driver_name, driver_orders in by_driver.items():
        admin_lines.append(f"*{driver_name}:*")
        for i, o in enumerate(driver_orders, 1):
            admin_lines.append(f"{i}. {o['customer_name']} — {o['site_address']}")
        admin_lines.append("")
    admin_ok = send_whatsapp_message(admin_phone, "\n".join(admin_lines))
    print(f"[Schedule] שליחה למנהל ({admin_phone}): {'✓' if admin_ok else '✗ נכשל'}")

    if not admin_ok:
        return {"ok": False, "orders_sent": 0,
                "error": f"וואטסאפ נכשל בשליחה למנהל ({admin_phone}) — בדוק WHATSAPP_ACCESS_TOKEN ומספר הטלפון בהגדרות"}

    # שלח לכל נהג את ההזמנות שלו
    driver_results = []
    with get_db() as conn:
        drivers_map = {d["id"]: dict(d) for d in conn.execute("SELECT * FROM drivers").fetchall()}

    for driver_name, driver_orders in by_driver.items():
        first = driver_orders[0]
        raw_phone    = _fmt_phone(first.get("driver_phone"))
        raw_personal = _fmt_phone(first.get("driver_personal_phone"))

        if not raw_phone and not raw_personal:
            driver_results.append({"name": driver_name, "ok": False, "reason": "אין טלפון"})
            continue

        header = (
            f"📋 *סידור יומי — {target_date}*\n"
            f"יש לך {len(driver_orders)} הזמנות:\n"
        )
        lines = []
        for i, o in enumerate(driver_orders, 1):
            lines.append(
                f"{i}. *{o['customer_name']}*\n"
                f"   📍 {o['site_address']}\n"
                f"   ⛽ {o['quantity']} ליטר"
                + (f"  🕐 {o['delivery_time']}" if o.get('delivery_time') else "")
                + (f"\n   👤 {o['contact_name']}" + (f" · {o['contact_phone']}" if o.get('contact_phone') else "") if o.get('contact_name') else "")
            )
        full_text = header + "\n".join(lines)

        # שלח טקסט רגיל לטלפון אישי (עדיף) או לטלפון הקבוצה
        target_phone = raw_personal or raw_phone
        ok_text = send_whatsapp_message(target_phone, full_text)

        # שלח כרטיסי ביצוע
        if raw_personal:
            # יש מספר אישי — כפתורים למספר האישי, טקסט לקבוצה
            for i, o in enumerate(driver_orders, 1):
                ok_card = send_order_card(raw_personal, o, i, len(driver_orders))
                if not ok_card:
                    print(f"[Schedule] ⚠ כרטיס נכשל לנהג {driver_name} ({raw_personal}) הזמנה #{o['id']}")
            if raw_phone and raw_phone != raw_personal:
                send_whatsapp_message(raw_phone, full_text)
        else:
            # אין מספר אישי — כפתורים ישירות לטלפון הנהג
            for i, o in enumerate(driver_orders, 1):
                ok_card = send_order_card(raw_phone, o, i, len(driver_orders))
                if not ok_card:
                    print(f"[Schedule] ⚠ כרטיס נכשל לנהג {driver_name} ({raw_phone}) הזמנה #{o['id']}")

        ok = ok_text
        phone_display = raw_personal or raw_phone
        driver_results.append({"name": driver_name, "ok": ok, "orders": len(driver_orders),
                                "phone": phone_display,
                                "reason": None if ok else "שגיאת וואטסאפ — בדוק מספר טלפון"})
        print(f"[Schedule] {driver_name} ({phone_display}): {'✓' if ok else '✗'}")

    return {"ok": True, "orders_sent": len(orders), "drivers": driver_results}


# ── Contact Suggestions ───────────────────────────────────────────────────────

@app.get("/api/contact-suggestions")
def get_contact_suggestions():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT contact_name, contact_phone,
                   COUNT(*) as usage_count
            FROM orders
            WHERE contact_name IS NOT NULL AND contact_name != ''
            GROUP BY contact_name, contact_phone
            ORDER BY usage_count DESC, contact_name
        """).fetchall()
    return [dict(r) for r in rows]


# ── Fleet Management ──────────────────────────────────────────────────────────

@app.get("/api/fleet/trucks")
def get_fleet_trucks():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT t.*, d.name as driver_name
            FROM fleet_trucks t LEFT JOIN drivers d ON t.driver_id = d.id
            ORDER BY t.name
        """).fetchall()
    return [dict(r) for r in rows]

@app.post("/api/fleet/trucks")
async def add_fleet_truck(request: Request):
    body = await request.json()
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO fleet_trucks (name, plate_number, driver_id, tanker_volume, notes) VALUES (?,?,?,?,?)",
            (body.get("name",""), body.get("plate_number",""),
             body.get("driver_id") or None, body.get("tanker_volume",""), body.get("notes",""))
        )
    return {"ok": True, "id": cur.lastrowid}

@app.put("/api/fleet/trucks/{tid}")
async def update_fleet_truck(tid: int, request: Request):
    body = await request.json()
    with get_db() as conn:
        conn.execute(
            "UPDATE fleet_trucks SET name=?, plate_number=?, driver_id=?, tanker_volume=?, notes=? WHERE id=?",
            (body.get("name",""), body.get("plate_number",""),
             body.get("driver_id") or None, body.get("tanker_volume",""), body.get("notes",""), tid)
        )
    return {"ok": True}

@app.delete("/api/fleet/trucks/{tid}")
def delete_fleet_truck(tid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM fleet_records WHERE truck_id=?", (tid,))
        conn.execute("DELETE FROM fleet_trucks WHERE id=?", (tid,))
    return {"ok": True}

@app.get("/api/fleet/records/{truck_id}")
def get_fleet_records(truck_id: int):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM fleet_records WHERE truck_id=? ORDER BY event_date DESC, created_at DESC",
            (truck_id,)
        ).fetchall()
    return [dict(r) for r in rows]

@app.post("/api/fleet/records")
async def add_fleet_record(request: Request):
    body = await request.json()
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO fleet_records (truck_id, category, title, event_date, expiry_date, cost, notes) VALUES (?,?,?,?,?,?,?)",
            (body.get("truck_id"), body.get("category",""), body.get("title",""),
             body.get("event_date",""), body.get("expiry_date",""),
             body.get("cost") or 0, body.get("notes",""))
        )
    return {"ok": True, "id": cur.lastrowid}

@app.put("/api/fleet/records/{rid}")
async def update_fleet_record(rid: int, request: Request):
    body = await request.json()
    with get_db() as conn:
        conn.execute(
            "UPDATE fleet_records SET category=?, title=?, event_date=?, expiry_date=?, cost=?, notes=? WHERE id=?",
            (body.get("category",""), body.get("title",""),
             body.get("event_date",""), body.get("expiry_date",""),
             body.get("cost") or 0, body.get("notes",""), rid)
        )
    return {"ok": True}

@app.delete("/api/fleet/records/{rid}")
def delete_fleet_record(rid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM fleet_records WHERE id=?", (rid,))
    return {"ok": True}


# ── WorkPlan persistence + bidirectional sync ────────────────────────────────

import json as _json
import urllib.request as _urllib
import threading as _threading
import time as _time

WORKPLAN_FILE = os.path.join(os.path.dirname(__file__), "workplan_data.json")
PROD_URL = "https://vezot-fuel.com"
_workplan_lock = _threading.Lock()


def _fetch_prod():
    """Fetch workplan from production server (server-to-server, no CORS)."""
    try:
        req = _urllib.Request(f"{PROD_URL}/api/workplan", headers={"Accept": "application/json"})
        with _urllib.urlopen(req, timeout=8) as resp:
            return _json.loads(resp.read().decode())
    except Exception:
        return None


def _push_prod(data: dict):
    """Push workplan to production server in a background thread."""
    def _do():
        try:
            payload = _json.dumps(data, ensure_ascii=False).encode()
            req = _urllib.Request(
                f"{PROD_URL}/api/workplan",
                data=payload,
                headers={"Content-Type": "application/json", "X-Sync-Source": "local"},
                method="POST",
            )
            _urllib.urlopen(req, timeout=8)
        except Exception:
            pass
    _threading.Thread(target=_do, daemon=True).start()


def _load_local():
    if not os.path.exists(WORKPLAN_FILE):
        return None
    try:
        with _workplan_lock:
            with open(WORKPLAN_FILE, "r", encoding="utf-8") as f:
                return _json.load(f)
    except Exception:
        return None


def _save_local(data: dict):
    with _workplan_lock:
        with open(WORKPLAN_FILE, "w", encoding="utf-8") as f:
            _json.dump(data, f, ensure_ascii=False, indent=2)


def _startup_sync():
    """On server start: always pull from production (source of truth).
    Local changes are pushed to production on every save, so production
    always has the most recent committed data."""
    prod = _fetch_prod()
    if not prod:
        print("[sync] production unreachable — using local data")
        return
    _save_local(prod)
    print("[sync] pulled data from production")


# Run startup sync in background so server starts immediately
_threading.Thread(target=_startup_sync, daemon=True).start()


@app.get("/api/workplan")
def get_workplan():
    data = _load_local()
    return JSONResponse(content=data)


@app.post("/api/workplan")
async def save_workplan(request: Request):
    from_sync = request.headers.get("X-Sync-Source")
    body = await request.json()
    if "lastModified" not in body:
        body["lastModified"] = int(_time.time() * 1000)
    _save_local(body)
    if not from_sync:          # don't re-push if this came from a sync (prevents infinite loop)
        _push_prod(body)
    return {"ok": True}


@app.post("/api/sync/pull")
def sync_pull():
    """Manually pull latest data from production."""
    prod = _fetch_prod()
    if not prod:
        return {"ok": False, "error": "לא ניתן להגיע לשרת הייצור"}
    local = _load_local()
    prod_ts  = prod.get("lastModified", 0)
    local_ts = local.get("lastModified", 0) if local else 0
    if prod_ts >= local_ts:
        _save_local(prod)
        return {"ok": True, "action": "pulled", "ts": prod_ts}
    return {"ok": True, "action": "already_newer", "ts": local_ts}


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
